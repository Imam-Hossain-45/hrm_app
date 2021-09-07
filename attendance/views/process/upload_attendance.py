from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin

from helpers import daily_record
from helpers.mixins import PermissionMixin
from django.conf import settings
from django.contrib import messages
from django.http import HttpResponseRedirect
from django.urls import reverse
from attendance.models import AttendanceData, AttendanceBreak, DailyRecord
from .manual_attendance import get_early_day_status, get_late_day_status, get_under_work
from datetime import datetime
import os
from helpers.functions import get_organizational_structure


class UploadAttendanceView(LoginRequiredMixin, PermissionMixin, TemplateView):
    """
        Upload attendance Data
    """
    login_url = settings.LOGIN_URL
    template_name = "attendance/process/manual_attendance/upload.html"
    permission_required = 'add_attendancedata'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['permissions'] = self.get_current_user_permission_list()
        context['org_items_list'] = get_organizational_structure()
        return context

    def set_csv_attendance_data(self, date, in_time, employee, out_date, out_time, break_time_list):
        created_id = ''
        if date not in ['', None] and in_time not in ['', None]:
            date = datetime.strptime(str(date), "%m/%d/%Y")
            in_time = datetime.strptime(in_time, '%H:%M')
            in_time = in_time.strftime('%H:%M')
            if out_time not in ['', None]:
                out_time = datetime.strptime(out_time, '%H:%M')
                out_time = out_time.strftime('%H:%M')
            if out_date not in ['', None]:
                out_date = datetime.strptime(str(out_date), "%m/%d/%Y")
            created = AttendanceData(employee_id=employee,
                                     date=date,
                                     in_time=in_time,
                                     out_date=out_date,
                                     out_time=out_time)

            try:
                created.save()
                created_id = created.pk
            except Exception as e:
                print(e)

            """Save data in ScheduleRecord and DailyRecord"""
            daily_record_data = {
                'employee': employee,
                'date': date
            }

            daily_record.set_daily_record(daily_record_data)

            """ get late or not """
            is_late, late_value, is_present, daily_pre_overtime_seconds = get_late_day_status(employee, date, in_time)
            is_early_out, early_out_value, daily_work_seconds, is_overtime, daily_post_overtime_seconds = \
                get_early_day_status(employee, date, in_time, out_time, out_date)
            daily_break_seconds = 0
            i = 0

            for j in range(0, 4):
                start_date = break_time_list[i]
                start_time = break_time_list[i + 1]
                end_date = break_time_list[i + 2]
                end_time = break_time_list[i + 3]
                i = i + 4
                if start_time not in ['', None]:
                    start_time = datetime.strptime(start_time, '%H:%M')
                    if start_date not in ['', None]:
                        start_date = datetime.strptime(str(start_date), "%m/%d/%Y")
                    if end_time not in ['', None]:
                        end_time = datetime.strptime(end_time, '%H:%M')
                    if end_date not in ['', None]:
                        end_date = datetime.strptime(str(end_date), "%m/%d/%Y")

                    update, created = AttendanceBreak.objects.update_or_create(attendance_id=created_id,
                                                                               break_start_date=start_date,
                                                                               break_start=start_time.strftime('%H:%M'),
                                                                               defaults={'break_end_date': end_date,
                                                                                         'break_end': end_time.strftime(
                                                                                             '%H:%M')})
                    if created:
                        if end_time not in ['', None] and end_date not in ['', None]:
                            # get daily break hour/seconds
                            daily_break_seconds = daily_break_seconds + (
                                datetime.combine(end_date, end_time.time()) - datetime.combine(start_date,
                                                                                               start_time.time())).seconds
            daily_work_seconds = daily_work_seconds - daily_break_seconds
            is_under_work, under_work_value = get_under_work(employee,
                                                             daily_work_seconds)

            """ Update dailyRecord """
            DailyRecordSave = DailyRecord.objects.filter(schedule_record__employee_id=employee,
                                                         schedule_record__date=date).update(
                daily_working_seconds=daily_work_seconds, is_overtime=is_overtime,
                daily_post_overtime_seconds=daily_post_overtime_seconds, late=is_late,
                late_value=late_value, early=is_early_out, early_out_value=early_out_value,
                under_work=is_under_work, under_work_value=under_work_value, is_present=is_present)

            if DailyRecordSave:
                print('Daily record found')
            else:
                print('No daily record found')

        return created_id

    def set_xls_attendance_data(self, date, in_time, employee, out_date, out_time, break_time_list):
        created_id = ''
        if date not in ['', None] and in_time not in ['', None]:
            created = AttendanceData(employee_id=employee,
                                     date=date.date(),
                                     in_time=in_time,
                                     out_date=out_date,
                                     out_time=out_time)
            try:
                created.save()
                created_id = created.pk
            except:
                print('error')

            """Save data in ScheduleRecord and DailyRecord"""
            daily_record_data = {
                'employee': employee,
                'date': date
            }

            daily_record.set_daily_record(daily_record_data)

            """ get late or not """
            is_late, late_value, is_present, daily_pre_overtime_seconds = \
                get_late_day_status(employee, date, in_time.strftime("%H:%M"))
            is_early_out, early_out_value, daily_work_seconds, is_overtime, daily_post_overtime_seconds = \
                get_early_day_status(employee, date, in_time.strftime("%H:%M"), out_time.strftime("%H:%M"), out_date)

            daily_break_seconds = 0
            i = 0
            for j in range(0, 4):
                start_date = break_time_list[i]
                start_time = break_time_list[i + 1]
                end_date = break_time_list[i + 2]
                end_time = break_time_list[i + 3]
                i = i + 4
                if start_time not in ['', None]:
                    update, created = AttendanceBreak.objects.update_or_create(attendance_id=created_id,
                                                                               break_start_date=start_date,
                                                                               break_start=start_time,
                                                                               defaults={'break_end_date': end_date,
                                                                                         'break_end': end_time})
                    if created:
                        if end_time not in ['', None] and end_date not in ['', None]:
                            # get daily break hour/seconds
                            daily_break_seconds = daily_break_seconds + (
                                datetime.combine(end_date, end_time) - datetime.combine(start_date, start_time)).seconds
            daily_work_seconds = daily_work_seconds - daily_break_seconds
            is_under_work, under_work_value = get_under_work(employee,
                                                             daily_work_seconds)

            """ Update dailyRecord """
            DailyRecordSave = DailyRecord.objects.filter(schedule_record__employee_id=employee,
                                                         schedule_record__date=date).update(
                daily_working_seconds=daily_work_seconds, is_overtime=is_overtime,
                daily_post_overtime_seconds=daily_post_overtime_seconds, late=is_late,
                late_value=late_value, early=is_early_out, early_out_value=early_out_value,
                under_work=is_under_work, under_work_value=under_work_value, is_present=is_present)

            if DailyRecordSave:
                print('Daily record found')
            else:
                print('No daily record found')

        return created_id

    def post(self, request, *args, **kwargs):
        success = True
        file_d = request.FILES['file']
        ext = os.path.splitext(file_d.name)[1]
        valid_extensions = ['.csv', '.xlsx', '.xls']
        if not ext.lower() in valid_extensions:
            messages.error(request, "File is not .xlsx, .xls, .csv type.")
            return HttpResponseRedirect(reverse("beehive_admin:attendance:upload_attendance"))
        i = 0
        if ext.lower() == '.csv':
            try:
                file_date = file_d.read().decode("utf-8")
                lines = file_date.split("\n")
                for line in lines:
                    i = i + 1
                    if line is not '':
                        if i == 1:
                            continue
                        fields = line.split(",")
                        data = {}
                        if fields[0] not in ['', None]:
                            data['employee'] = fields[0]
                            data['date'] = fields[1]
                            data['in_time'] = fields[2]
                            if fields[3] is not '':
                                data['out_date'] = fields[3]
                            else:
                                data['out_date'] = fields[1]
                            data['out_time'] = fields[4]

                            break_time_list = []
                            for d in range(5, 25):
                                break_time_list.append(fields[d])
                            success = self.set_csv_attendance_data(data['date'], data['in_time'], data['employee'],
                                                                   data['out_date'], data['out_time'], break_time_list)
                if success:
                    messages.success(request, 'Successfully uploaded.')
                else:
                    messages.error(request, "Error in {} no line.".format(i))
                # else:
                #     form.errors.as_json()
                #     messages.error(request, form.errors)
            except Exception as e:
                messages.error(request, "Error in {} no line.".format(i))
        else:
            try:
                # upload xslx or xsl sheet data
                from openpyxl import load_workbook
                wb = load_workbook(file_d)
                sheet = wb.active
                for row in sheet.iter_rows():
                    i = i + 1
                    if i == 1:
                        continue
                    j = 0
                    my_dict = {}
                    for cell in row:
                        my_dict[j] = cell.value
                        j = j + 1
                    data = {}
                    data['employee'] = my_dict[0]
                    data['date'] = my_dict[1] if my_dict[1] is not None else my_dict[1]
                    data['in_time'] = my_dict[2]
                    if my_dict[3] is not '':
                        data['out_date'] = my_dict[3] if my_dict[3] is not None else my_dict[3]
                    else:
                        data['out_date'] = data['date']
                    data['out_time'] = my_dict[4]

                    break_time_list = []
                    for d in range(5, 25):
                        break_time_list.append(my_dict[d])

                    success = self.set_xls_attendance_data(data['date'], data['in_time'], data['employee'],
                                                           data['out_date'], data['out_time'], break_time_list)
                if success:
                    messages.success(request, 'Successfully uploaded.')
                else:
                    messages.error(request, "Error in {} no line.".format(i))
            except Exception as e:
                messages.error(request, "Error in {} no line.".format(i))

        return HttpResponseRedirect(reverse('beehive_admin:attendance:upload_attendance'))
